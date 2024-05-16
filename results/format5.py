from openpyxl import Workbook
import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles import Font, Alignment, Border, Side
import os
import uuid


class f5:

    def __init__(self, reqdata):
        self.creditlist = []
        self.schooldata = reqdata["student-data"]
        self.schooldata['Enrollment No'] = pd.to_numeric(
            self.schooldata['Enrollment No'])
        series = self.schooldata['Enrollment No']
        for i in range(len(series)):
            series[i] = f"{int(series[i]):110d}"
        self.schooldata['Enrollment No'] = series
        print(self.schooldata)
        self.course = reqdata["course"]
        self.shift = reqdata["shift"]
        self.passout_year = reqdata["passout_year"]
        self.file_name = str(uuid.uuid4())
        self.path = os.path.join(os.path.dirname(
            __file__), "buffer_files", f"{self.file_name}.xlsx")
        self.sems = len(reqdata["data"].values())
        self.dataframes = [i for i in reqdata["data"].values()]

        for d in self.dataframes:
            credit = round(float(d.iloc[2, -4])/float(d.iloc[2, -3]))
            self.creditlist.append(credit)
        self.total_credits = sum(self.creditlist)

        for i, d in enumerate(self.dataframes):
            series = d['Enrollment Number']
            for j in range(len(series)):
                if j == 0:
                    continue
                print(series[j])
                series[j] = f"{int(series[j]):110d}"
            d['Enrollment Number'] = series
            print(d)
        self.no_of_students = len(self.dataframes[0])
        self.sem_map = {
            1: 'I',
            2: 'II',
            3: 'III',
            4: 'IV',
            5: 'V',
            6: 'VI',
            7: 'VII',
            8: 'VIII',
            9: 'IX',
            10: 'X'
        }

    def format(self):

        wb = Workbook()
        ws = wb.active
        for i in range(1, 8+self.sems*2+5):
            for j in range(self.no_of_students):
                ws.cell(row=8+j, column=self.sems*2 + 10).value = ""
        ws['A1'] = 'Maharaja Surajmal Institute'
        ws['A2'] = 'Department of Computer Application'
        ws['A3'] = 'SECOND SHIFT'
        ws['A4'] = 'BCA Section-'
        ws['A5'] = 'Batch: 2017-20'
        for i in range(1, 6):
            ws.cell(row=i, column=1).alignment = Alignment(horizontal='center')
            ws.merge_cells(start_row=i, start_column=1,
                           end_row=i, end_column=12)
        ws['A8'] = 'S.No.'
        ws['B8'] = 'Enrollment No.'
        ws['C8'] = 'Name of the Student'
        ws['D8'] = 'Section'
        ws['E8'] = 'X %'
        ws['F8'] = 'XII %'
        ws['G8'] = 'Aggregate % of X and XII'
        ws['H8'] = 'Category'
        for i in range(1, self.sems*2+1, 2):
            ws.cell(row=8, column=8+i).value = f'{self.sem_map[(i+1)//2]} Sem%'
            ws.cell(row=8, column=9+i).value = 'Category'
        for i in range(self.schooldata.shape[0]):
            ws.cell(row=9+i, column=1).value = i+1
            ws.cell(row=9+i, column=2).value = self.schooldata.iloc[i, 1]
            ws.cell(row=9+i, column=3).value = self.schooldata.iloc[i, 0]
            ws.cell(row=9+i, column=4).value = "A"
            ws.cell(row=9+i, column=5).value = self.schooldata.iloc[i, 2]
            ws.cell(row=9+i, column=6).value = self.schooldata.iloc[i, 3]
            ws.cell(row=9+i, column=7).value = (
                float(self.schooldata.iloc[i, 2]) + float(self.schooldata.iloc[i, 3]))/2
            ws.cell(row=9+i, column=8).value = "O" if ws.cell(row=9+i, column=6).value >= 90 else "A" if ws.cell(
                row=9+i, column=6).value >= 75 else "B" if ws.cell(row=9+i, column=6).value >= 60 else "C" if ws.cell(row=9+i, column=6).value >= 50 else "D" if ws.cell(row=9+i, column=6).value >= 40 else "Fail"
            for j in range(1, self.sems+1):
                sem = self.dataframes[j-1]
                total_marks = 0
                if ws.cell(row=9+i, column=2).value in sem['Enrollment Number'].values:
                    reappear_codes = absent_codes = ""

                    marks = sem[sem['Enrollment Number']
                                == ws.cell(row=9+i, column=2).value]
                    ws.cell(row=9+i, column=7+j *
                            2).value = marks['CGPA%'].values[0]
                    ws.cell(row=9+i, column=8+j*2).value = "O" if marks['CGPA%'].values[0] >= 90 else "A" if marks['CGPA%'].values[0] >= 75 else "B" if marks[
                        'CGPA%'].values[0] >= 60 else "C" if marks['CGPA%'].values[0] >= 50 else "D" if marks['CGPA%'].values[0] >= 40 else "Fail"
                    ws.cell(row=9+i, column=4).value = marks['Sec'].values[0]

                    total_credits = sum(self.creditlist)

                    total_marks += marks['Total'].values[0]
                    ws.cell(row=9+i, column=8+self.sems*2+1).value = total_marks/total_credits if ws.cell(
                        row=9+i, column=8+self.sems*2+1).value == None else ws.cell(row=9+i, column=8+self.sems*2+1).value + total_marks/total_credits
                    ws.cell(row=9+i, column=10+self.sems*2).value = "O" if ws.cell(row=9+i, column=9+self.sems*2).value >= 90 else "A" if ws.cell(row=9+i, column=9+self.sems*2).value >= 75 else "B" if ws.cell(
                        row=9+i, column=9+self.sems*2).value >= 60 else "C" if ws.cell(row=9+i, column=9+self.sems*2).value >= 50 else "D" if ws.cell(row=9+i, column=9+self.sems*2).value >= 40 else "Fail"
                    reappear_codes += ",".join(
                        marks.iloc[:, -2].values[0].split(",")) if marks.iloc[:, -2].values[0] != None else ""
                    ws.cell(row=9+i, column=11+self.sems*2).value = ws.cell(row=9 +
                                                                            i, column=8+self.sems*2+3).value + reappear_codes if ws.cell(row=9+i, column=8+self.sems*2+3).value != None else reappear_codes
                    absent_codes = ",".join(
                        marks.iloc[:, -1].values[0].split(",")) if marks.iloc[:, -1].values[0] != None else ""
                    ws.cell(row=9+i, column=12+self.sems*2).value = ws.cell(row=9 +
                                                                            i, column=8+self.sems*2+4).value + absent_codes if ws.cell(row=9+i, column=8+self.sems*2+4).value != None else absent_codes

                else:

                    ws.cell(row=9+i, column=7+j*2).value = "NA"
                    ws.cell(row=9+i, column=8+j*2).value = "NA"

        # designing the borders
        for i in range(1, 8+self.sems*2+5):
            for j in range(self.no_of_students):
                ws.cell(row=8+j, column=i).border = Border(
                    top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        ptr = self.sems*2+8
        ws.cell(row=8, column=ptr +
                1).value = f'Aggregate of {str(",".join(self.sem_map[i] for i in range(1,self.sems+1)))} Semesters'
        ws.cell(row=8, column=ptr+2).value = 'Category'
        ws.cell(row=8, column=ptr+3).value = 'Reappear Paper Codes'
        ws.cell(row=8, column=ptr+4).value = 'Absent Paper Codes'
        ptr += 6
        ws.cell(row=8, column=ptr).value = 'Percentage'
        ws.cell(row=8, column=ptr+1).value = 'category'
        for i in range(ptr, ptr+2):
            for j in range(7):
                ws.cell(row=8+j, column=i).border = Border(
                    top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        ws.cell(row=9, column=ptr).value = '>=90%'
        ws.cell(row=10, column=ptr).value = '75 to 89.99%'
        ws.cell(row=11, column=ptr).value = '60 to 74.99%'
        ws.cell(row=12, column=ptr).value = '50 to 59.99%'
        ws.cell(row=13, column=ptr).value = '40 to 49.99%'
        ws.cell(row=14, column=ptr).value = '<40%'
        ws.cell(row=9, column=ptr+1).value = 'O'
        ws.cell(row=10, column=ptr+1).value = 'A'
        ws.cell(row=11, column=ptr+1).value = 'B'
        ws.cell(row=12, column=ptr+1).value = 'C'
        ws.cell(row=13, column=ptr+1).value = 'D'
        ws.cell(row=14, column=ptr+1).value = 'Fail'
        start_row = 8
        end_row = 14
        end_col = ptr+1
        for i in range(start_row, end_row+1):
            for j in range(ptr, end_col+1):
                ws.cell(row=i, column=j).font = Font(color="FF0000")
        start_row = 12+self.no_of_students
        end_row = 12+self.no_of_students+5
        start_col = 3
        end_col = 4
        ws.cell(row=start_row-1, column=start_col).value = 'Summary Based on Columns'
        for i in range(start_row, end_row+1):
            for j in range(start_col, end_col+1):
                ws.cell(row=i, column=j).border = Border(
                    top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
        ws.cell(row=12+self.no_of_students, column=3).value = 'Category '
        ws.cell(row=12+self.no_of_students, column=4).value = 'No. of Students'
        ws.cell(row=12+self.no_of_students+1, column=3).value = 'O'
        ws.cell(row=12+self.no_of_students+2, column=3).value = 'A'
        ws.cell(row=12+self.no_of_students+3, column=3).value = 'B'
        ws.cell(row=12+self.no_of_students+4, column=3).value = 'C'
        ws.cell(row=12+self.no_of_students+5, column=3).value = 'D'
        ws.cell(row=12+self.no_of_students+6, column=3).value = 'Fail'
        ws.cell(row=12+self.no_of_students+7,
                column=3).value = 'No. of Absentees'

        # strip the data
        for i in range(1, 8+self.sems*2+5):
            for j in range(self.no_of_students):
                ws.cell(row=8+j, column=i).value = str(ws.cell(row=8 +
                                                               j, column=i).value).strip()
        wb.save(self.path)
        return f"{self.file_name}.xlsx"
