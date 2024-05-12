from openpyxl import load_workbook
import pandas as pd
from openpyxl.styles import Font
from openpyxl.styles import Font, Alignment, Border, Side
import os


class ResultProcessor:



    subject_name_mapping = {}


    def __init__(self, input_file, output_file , subject_name_mapping, exclude_subject_dict , footers_to_add , headers_to_add,credits_mapping):
        if bool(exclude_subject_dict):
            self.exclude_subject_code = list(exclude_subject_dict.keys())[0]
            self.exclude_columns=[exclude_subject_dict[self.exclude_subject_code]]
        else:
            self.exclude_subject_code,self.exclude_columns= "",[]
        
        self.subject_name_mapping = subject_name_mapping
        self.input_file = input_file
        self.output_file = output_file
        self.df = None
        self.subject_column_renaming = {}
        self.footers_to_add = footers_to_add
        self.headers_to_add = headers_to_add
        self.total_subjects = len(subject_name_mapping)
        self.credits_mapping=credits_mapping



    def read_data(self):
        if hasattr(self.input_file, 'read'):
            # Reset the file pointer to the beginning of the file
            self.input_file.seek(0)
            self.df = pd.read_csv(self.input_file)
        else:
            # Assume input_file is a file path
            self.df = pd.read_csv(self.input_file)


    def rename_columns(self):
        

        for name in self.df.columns:
            name.split('/')[0].strip()
        self.df.rename(columns=self.subject_name_mapping , inplace=True)
        new_df=self.df.iloc[:,10:15]
        
        

    def calculate_total(self):
        
        
        self.df['Total'] = 0
        
        credits_mapping=self.credits_mapping

        
        for subject, credits in credits_mapping.items():
            
            self.df['Total'] += self.df.apply(lambda row: (float(row[subject]) * credits)
                                             if row[subject] and str(row[subject]).strip().isdigit() else 0, axis=1)
        
    def calculate_cgpa(self):
        
        self.df['CGPA%'] = 0.0
        credits_mapping=self.credits_mapping

        for subject, credits in credits_mapping.items():
            self.df['CGPA%'] += self.df.apply(lambda row: (float(row[subject]) * credits)
                                             if row[subject] and str(row[subject]).strip().isdigit() else 0, axis=1)

        total_credits = sum(credits_mapping.values())
        self.df['CGPA%'] /= total_credits.__round__(4)

    def process_reappear(self):
        self.df['Reapper Paper Codes'] = ''

        def is_numeric(value):
            try:
                int(value)
                return True
            except ValueError:
                return False

        def filter_reappear(row):
            
            # Extract reappear subjects based on condition
            reappear_subjects = [
                col.split('(')[0].strip()
                for col in self.df.columns[4:-4]
                if 'Total' in col
                and is_numeric(row[col]) and 1 <= int(row[col]) < 40 
                and not print(col.split('(')[0].strip()) # Check numeric and range
                and col not in self.exclude_columns
                and self.exclude_subject_code not in col
                and row[col] != '0'
                and not any(row[col] == 'Absent Paper Codes' for col in self.df.columns[4:-2] if 'External' in col and col not in self.exclude_columns)
            ]
            print(reappear_subjects)

            # Return unique reappear subjects as a comma-separated string
            return ','.join(set(reappear_subjects))

        self.df['Reapper Paper Codes'] = self.df.apply(filter_reappear, axis=1)
        self.df['Reapper Paper Codes'] = self.df['Reapper Paper Codes'].apply(
            lambda x: ','.join([s[3:6] for s in x.split(',')]))

    def process_absents(self):
        self.df['Absent Paper Codes'] = ''

        def filter_absent(row):

            return ','.join(set([col.split('(')[0].strip() for col in self.df.columns[4:-4]
                                 if 'External' in col and str(row[col]).strip() == '0' and col not in self.exclude_columns
                                 ]))

        self.df['Absent Paper Codes'] = self.df.apply(filter_absent, axis=1)
        self.df['Absent Paper Codes'] = self.df['Absent Paper Codes'].apply(
            lambda x: ','.join([s[3:6] for s in x.split(',')]))
        self.df.to_csv(os.path.join(os.path.dirname(__file__), "buffer_files", "temp.csv"))


    def update_reappear_absent_columns(self):
        reappear_lists = self.df['Reapper Paper Codes'].apply(
            lambda x: x.split(',') if pd.notna(x) else [])
        absent_lists = self.df['Absent Paper Codes'].apply(
            lambda x: x.split(',') if pd.notna(x) else [])

        common_subjects = set.intersection(
            *map(set, reappear_lists.tolist() + absent_lists.tolist()))

        def update_columns(row, column_name):
            return ','.join([subject for subject in row if subject not in common_subjects])

        self.df['Reapper Paper Codes'] = reappear_lists.apply(
            lambda row: update_columns(row, 'Reapper Paper Codes'))
        self.df['Absent Paper Codes'] = absent_lists.apply(
            lambda row: update_columns(row, 'Absent Paper Codes'))

    def final_rename_columns(self):
        self.temp = self.subject_name_mapping
        credits_mapping=self.credits_mapping
        for key , value in self.temp.items():
            self.subject_column_renaming[value] = value

        for key , value in self.subject_column_renaming.items():

            if "internal" in value.lower():

                value = value.replace("Internal",str(credits_mapping[key.replace("Internal" , "Total")]))

                
                self.subject_column_renaming[key] = value
            else:
                value = ""
                
        
        self.df.rename(columns=self.subject_column_renaming, inplace=True)
    
    def initialize_for_elective_df(self ,df):
        self.df = df
    def save_result(self):
        output_file_path = os.path.join(os.path.dirname(__file__), "buffer_files", self.output_file)
        writer = pd.ExcelWriter(output_file_path, engine='xlsxwriter', engine_kwargs={'options': {'encoding': 'utf-8'}})

        self.df.to_excel(writer, sheet_name='ResultSheet', index=False)
        worksheet = writer.sheets['ResultSheet']

        fixed_column_width = 4
        left_aligned_format = writer.book.add_format({'align': 'left'})

        worksheet.set_column(3, 30, fixed_column_width, left_aligned_format)
        worksheet.set_column(0, 0,4,left_aligned_format)
        worksheet.set_column(1, 1,10,left_aligned_format)
        worksheet.set_column(2, 2,20,left_aligned_format)
        worksheet.set_column(31, 31,5,left_aligned_format)
        worksheet.set_column(32, 32,7,left_aligned_format)
        writer._save()
        
        output_file_path = os.path.join(os.path.dirname(__file__), "buffer_files", self.output_file)
        workbook = load_workbook(output_file_path)


      
        sheet = workbook.active
        start_row = sheet.max_row + 3

        for i, footer_row in enumerate(self.footers_to_add):
            for j, footer_value in enumerate(footer_row, start=1):
                sheet.cell(row=start_row + i, column=j, value=footer_value)
        sheet.insert_rows(1, amount=len(self.headers_to_add) + 1)

        bold_font = Font(name='Times New Roman', bold=True)
        middle_alignment = Alignment(horizontal='center', vertical='center')

        for i, header_row in enumerate(self.headers_to_add):
            for j, header_value in enumerate(header_row, start=1):
                cell = sheet.cell(row=i + 1, column=j, value=header_value)
                cell.font = bold_font
                cell.alignment = middle_alignment

        for i in range(1, 4):
            
            sheet.merge_cells(start_row=i, start_column=1,
                              end_row=i, end_column=len(self.subject_name_mapping)+3)
        start_row = 6
        end_row = 6
        start_column = 1
        end_column = 7+len(self.subject_name_mapping)*3
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    cell.value = cell.value.replace('Internal','Int')
                    cell.value = cell.value.replace('External','Ext')
        sheet.insert_rows(7, amount=1)
        num=(self.df.shape[0]-4)
        result = ''
        while num > 0:
            num, remainder = divmod(num - 1, 26)
            result = chr(65 + remainder) + result
        incremented_column = lambda col: col[:-1] + chr(ord(col[-1]) + 1)
        numplus1=incremented_column(result)
        sheet.move_range(result+"5:"+result+"6", rows=1, cols=0)
        sheet.move_range(numplus1+"5:"+numplus1+"6", rows=1, cols=0)

        start_row = 5
        end_row = 6+self.df.shape[0]
        start_column = 1
        end_column = 8+len(self.subject_name_mapping)
        style=Side(style='thin')
        thin_border = Border(left=style,
                             right=style,
                             top=style,
                             bottom=style)
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
        start_row = 6
        end_row = 6+self.df.shape[0]
        start_column = 4
        end_column = 4
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = Font(bold=True)
        start_row = 7
        end_row = 7
        start_column = 1
        end_column = 7+len(self.subject_name_mapping)
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell.value = ''
        start_row = 1
        end_row = 6+self.df.shape[0]
        start_column = 1
        end_column = 8+len(self.subject_name_mapping)
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.font.bold:
                    cell.font = Font(name='Times New Roman', bold=True)
                else:
                    cell.font = Font(name='Times New Roman')
        sheet.row_dimensions[5].height = 60
        for col in sheet.columns:
            for cell in col:
                if cell.row == 5:
                    cell.alignment = Alignment(wrap_text=True,horizontal='center', vertical='bottom')
                    cell.font = Font(name='Times New Roman', bold=True)
        start_row = 6
        end_row = 6
        start_column = 1
        end_column = 8+len(self.subject_name_mapping)
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = Font(name='Times New Roman', bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        start_row = 5
        end_row = 5
        start_column = 1
        end_column = len(self.subject_name_mapping)
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column ):
                cell = sheet.cell(row=row, column=col)
                cell.font = Font(name='Times New Roman', bold=True)
                cell.alignment = Alignment(wrap_text=True)
        
        for i in range(int(len(self.subject_name_mapping)/3)):
            start_col = 5 + 3 * i
            end_col = 7 + 3 * i
            sheet.merge_cells(start_row=5, start_column=start_col,
                              end_row=5, end_column=end_col)
        for i in range(2):
            col = 5+len(self.subject_name_mapping)+i

            sheet.merge_cells(start_row=5, start_column=col,
                              end_row=6, end_column=col)
        output_file_path = os.path.join(os.path.dirname(__file__), "buffer_files", self.output_file)

        workbook.save(output_file_path)
        
        return True , self.df


