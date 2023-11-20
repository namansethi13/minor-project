import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side


class ResultProcessor:
    exclude_columns = ['SAUE (Internal)']
    exclude_subject_code = "20136"
    subject_column_renaming = {'020102 Applied Maths (Internal)': '20102 Applied Maths(4)',
                               '020102 Applied Maths (External)': '',
                               '020102 Applied Maths (Total)': '',
                               '020104 Web Based Programming (Internal)': '20104 Web Based Programming(4)',
                               '020104 Web Based Programming (External)': '',
                               '020104 Web Based Programming (Total)': '',
                               '020106 Data Structures & Algorithm Using C (Internal)': '20106 Data Structures & Algorithm Using C (4)',
                               '020106 Data Structures & Algorithm Using C (External)': '',
                               '020106 Data Structures & Algorithm Using C (Total)': '',
                               '020108 DBMS (Internal)': '20108 DBMS (4)',
                               '020108 DBMS (External)': '',
                               '020108 DBMS (Total)': '',
                               '020110 EVS (Internal)': '20110 EVS (2)',
                               '020110 EVS (External)': '',
                               '020110 EVS (Total)': '',
                               '020136 SAUE (Internal)': '20136 SAUE (2)',
                               '020136 SAUE (External)': '',
                               '020136 SAUE (Total)': '',
                               '020172 Practical IV-WBP Lab (Internal)': '20172 Practical IV-WBP Lab (2)',
                               '020172 Practical IV-WBP Lab (External)': '',
                               '020172 Practical IV-WBP Lab (Total)': '',
                               '020174 Practical- V DS Lab (Internal)': '20174 Practical- V DS Lab (2)',
                               '020174 Practical- V DS Lab (External)': '',
                               '020174 Practical- V DS Lab (Total)': '',
                               '020176 Practical- VI DBMS Lab (Internal)': '20176 Practical- VI DBMS Lab (2)',
                               '020176 Practical- VI DBMS Lab (External)': '',
                               '020176 Practical- VI DBMS Lab (Total)': ''
                               }
    subject_name_mapping = {'020102(4)': '020102 Applied Maths (Internal)',
                            '020102(4).1': '020102 Applied Maths (External)',
                            '020102(4).2': '020102 Applied Maths (Total)',
                            '020104(4)': '020104 Web Based Programming (Internal)',
                            '020104(4).1': '020104 Web Based Programming (External)',
                            '020104(4).2': '020104 Web Based Programming (Total)',
                            '020106(4)': '020106 Data Structures & Algorithm Using C (Internal)',
                            '020106(4).1': '020106 Data Structures & Algorithm Using C (External)',
                            '020106(4).2': '020106 Data Structures & Algorithm Using C (Total)',
                            '020108(4)': '020108 DBMS (Internal)',
                            '020108(4).1': '020108 DBMS (External)',
                            '020108(4).2': '020108 DBMS (Total)',
                            '020110(2)': '020110 EVS (Internal)',
                            '020110(2).1': '020110 EVS (External)',
                            '020110(2).2': '020110 EVS (Total)',
                            '020136(2)': '020136 SAUE (Internal)',
                            '020136(2).1': '020136 SAUE (External)',
                            '020136(2).2': '020136 SAUE (Total)',
                            '020172(2)': '020172 Practical IV-WBP Lab (Internal)',
                            '020172(2).1': '020172 Practical IV-WBP Lab (External)',
                            '020172(2).2': '020172 Practical IV-WBP Lab (Total)',
                            '020174(2)': '020174 Practical- V DS Lab (Internal)',
                            '020174(2).1': '020174 Practical- V DS Lab (External)',
                            '020174(2).2': '020174 Practical- V DS Lab (Total)',
                            '020176(2)': '020176 Practical- VI DBMS Lab (Internal)',
                            '020176(2).1': '020176 Practical- VI DBMS Lab (External)',
                            '020176(2).2': '020176 Practical- VI DBMS Lab (Total)',
                            }
    headers_to_add = [["Maharaja Surajmal Institute"],
                      ["BCA(M) Batch 2022-2025"],
                      ["Class-: BCA  II Semester Batch [2022-2025]     Jan- June 2023"],
                      ]
    footers_to_add = [["102-Applied Maths Dr. Anchal Tehlan (Sec A & B)"],
                      ["104-WBP - Mr. Sundeep Kumar(A) & Ms. Kanika Kundu (B)"],
                      ["106-Data Struc Using C - Dr.Neetu Anand(A )   Mr.Manoj Kumar (B )"],
                      ["108-DBMS - Ms.Kanika Kundu (A) & Ms.Vinita Tomar(B)"],
                      ["110-EVS - Dr.Manju Dhillon (Sec A & Sec B)"],
                      ["172-WBP Lab - Mr.Sundeep Kumar/ Dr.Neetu Narwal (Sec A) &  Ms.Kanika Kundu(Sec A) & Dr.Neetu Narwal (Sec B)"],
                      ["174- DS Lab - Dr.Neetu Anand /Dr.Kumar Gaurav (A ) &  Mr.Manoj Kumar (B )"],
                      ["176- DBMS Lab - Ms.Kanika Kundu / Mr. Siddharth Shankar (A)   &  Ms.Vinita Tomar (B)"],
                      [""],
                      ["Class Coordinator: Ms.Anchal Tehlan (Sec A) - Mr.Manoj Kumar (Sec B)"],
                      ]

    def __init__(self, input_file, output_file):
        self.input_file = input_file
        self.output_file = output_file
        self.df = None

    def read_data(self):
        self.fd = pd.read_csv(self.input_file)

    def rename_columns(self):
        self.df.rename(columns=self.subject_name_mapping, inplace=True)

    def process_reappear(self):
        self.df['Reapper Paper Codes'] = ''

        def is_numeric(value):
            try:
                int(value)
                return True
            except ValueError:
                return False

        def filter_reappear(row):
            return ','.join(set([col.split('(')[0].strip() for col in self.df.columns[4:-4]
                                 if 'Total' in col and is_numeric(row[col]) and int(row[col].strip()) < 40 and int(row[col].strip()) >= 1
                                 and col not in self.exclude_columns
                                 and self.exclude_subject_code not in col
                                 and not any(row[col].strip() == 'Absent Paper Codes' for col in self.df.columns[4:-2]
                                             if 'External' in col and col not in self.exclude_columns
                                             and self.exclude_subject_code not in col) and row[col] != '0']))

        self.df['Reapper Paper Codes'] = self.df.apply(filter_reappear, axis=1)
        self.df['Reapper Paper Codes'] = self.df['Reapper Paper Codes'].apply(
            lambda x: ','.join([s[3:6] for s in x.split(',')]))

    def process_absents(self):
        self.df['Absent Paper Codes'] = ''

        def filter_absent(row):
            return ','.join(set([col.split('(')[0].strip() for col in self.df.columns[4:-2]
                                 if 'External' in col and row[col].strip() == '0' and col not in self.exclude_columns
                                 and self.exclude_subject_code not in col]))

        self.df['Absent Paper Codes'] = self.df.apply(filter_absent, axis=1)
        self.df['Absent Paper Codes'] = self.df['Absent Paper Codes'].apply(
            lambda x: ','.join([s[3:6] for s in x.split(',')]))

    def calculate_cgpa(self):
        self.df['CGPA%'] = 0.0

        credits_mapping = {'020102 Applied Maths (Total)': 4,
                           '020104 Web Based Programming (Total)': 4,
                           '020106 Data Structures & Algorithm Using C (Total)': 4,
                           '020108 DBMS (Total)': 4,
                           '020110 EVS (Total)': 2,
                           '020136 SAUE (Total)': 2,
                           '020172 Practical IV-WBP Lab (Total)': 2,
                           '020174 Practical- V DS Lab (Total)': 2,
                           '020176 Practical- VI DBMS Lab (Total)': 2,
                           }

        for subject, credits in credits_mapping.items():
            self.df['CGPA%'] += self.df.apply(lambda row: (float(row[subject]) * credits)
                                             if row[subject] and str(row[subject]).strip().isdigit() else 0, axis=1)

        total_credits = sum(credits_mapping.values())
        self.df['CGPA%'] /= total_credits.__round__(4)
    def calculate_total(self):
        self.df['Total'] = 0

        credits_mapping = {'020102 Applied Maths (Total)': 4,
                           '020104 Web Based Programming (Total)': 4,
                           '020106 Data Structures & Algorithm Using C (Total)': 4,
                           '020108 DBMS (Total)': 4,
                           '020110 EVS (Total)': 2,
                           '020136 SAUE (Total)': 2,
                           '020172 Practical IV-WBP Lab (Total)': 2,
                           '020174 Practical- V DS Lab (Total)': 2,
                           '020176 Practical- VI DBMS Lab (Total)': 2,
                           }

        for subject, credits in credits_mapping.items():
            self.df['Total'] += self.df.apply(lambda row: (float(row[subject]) * credits)
                                             if row[subject] and str(row[subject]).strip().isdigit() else 0, axis=1)


    def save_result(self):

        writer = pd.ExcelWriter(self.output_file, engine='xlsxwriter')

        self.df.to_excel(writer, sheet_name='ResultSheet', index=False)

        worksheet = writer.sheets['ResultSheet']

        fixed_column_width = 4
        left_aligned_format = writer.book.add_format({'align': 'left'})

        worksheet.set_column(3, 30, fixed_column_width, left_aligned_format)
        worksheet.set_column(0, 0,4,left_aligned_format)
        worksheet.set_column(1, 1,10,left_aligned_format)
        worksheet.set_column(2, 2,20,left_aligned_format)
        worksheet.set_column(31, 31,5,left_aligned_format)
        worksheet.set_column(32, 32,7,left_aligned_format)
        writer._save()

    from openpyxl.styles import Font

    def add_headers_to_excel(self):
        workbook = load_workbook(self.output_file)
        sheet = workbook.active
        sheet.insert_rows(1, amount=len(self.headers_to_add) + 1)

        bold_font = Font(name='Times New Roman', bold=True)
        middle_alignment = Alignment(horizontal='center', vertical='center')

        for i, header_row in enumerate(self.headers_to_add):
            for j, header_value in enumerate(header_row, start=1):
                cell = sheet.cell(row=i + 1, column=j, value=header_value)
                cell.font = bold_font
                cell.alignment = middle_alignment

        for i in range(1, 4):
            sheet.merge_cells(start_row=i, start_column=1,
                              end_row=i, end_column=30)

        

        workbook.save(self.output_file)

    def update_reappear_absent_columns(self):
        reappear_lists = self.df['Reapper Paper Codes'].apply(
            lambda x: x.split(',') if pd.notna(x) else [])
        absent_lists = self.df['Absent Paper Codes'].apply(
            lambda x: x.split(',') if pd.notna(x) else [])

        common_subjects = set.intersection(
            *map(set, reappear_lists.tolist() + absent_lists.tolist()))

        def update_columns(row, column_name):
            return ','.join([subject for subject in row if subject not in common_subjects])

        self.df['Reapper Paper Codes'] = reappear_lists.apply(
            lambda row: update_columns(row, 'Reapper Paper Codes'))
        self.df['Absent Paper Codes'] = absent_lists.apply(
            lambda row: update_columns(row, 'Absent Paper Codes'))

    def add_footers_to_excel(self):
        workbook = load_workbook(self.output_file)
        sheet = workbook.active
        start_row = sheet.max_row + 3

        for i, footer_row in enumerate(self.footers_to_add):
            for j, footer_value in enumerate(footer_row, start=1):
                sheet.cell(row=start_row + i, column=j, value=footer_value)

        workbook.save(self.output_file)

    def final_rename_columns(self):
        self.df.rename(columns=self.subject_column_renaming, inplace=True)

    def merge_column_names(self):
        workbook = load_workbook(self.output_file)
        sheet = workbook.active
        for i in range(1, 10):
            start_col = 3 * i + 2
            end_col = 3 * i + 4
            sheet.merge_cells(start_row=5, start_column=start_col,
                              end_row=5, end_column=end_col)
        for i in range(2):
            col = 34+i
            sheet.merge_cells(start_row=5, start_column=col,
                              end_row=6, end_column=col)
        workbook.save(self.output_file)

    def add_borders_to_data(self):
        style=Side(style='thin')
        thin_border = Border(left=style,
                             right=style,
                             top=style,
                             bottom=style)
        workbook = load_workbook(self.output_file)
        worksheet = workbook.active
        start_row = 5
        end_row = 118
        start_column = 1
        end_column = 35
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
        workbook.save(self.output_file)
    def styling_sixth_row(self):
        workbook = load_workbook(self.output_file)
        worksheet = workbook.active
        start_row = 6
        end_row = 6
        start_column = 1
        end_column = 35
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = Font(name='Times New Roman', bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        workbook.save(self.output_file)

    def set_height_fifth_row(self):
        workbook = load_workbook(self.output_file)
        worksheet = workbook.active
        worksheet.row_dimensions[5].height = 60
        for col in worksheet.columns:
            for cell in col:
                if cell.row == 5:
                    cell.alignment = Alignment(wrap_text=True,horizontal='center', vertical='bottom')
                    cell.font = Font(name='Times New Roman', bold=True)
                    
        workbook.save(self.output_file)
        
    def replace(self):
        #replace internal and extenal with int and ext in 6th row
        workbook = load_workbook(self.output_file)
        worksheet = workbook.active 
        start_row = 6
        end_row = 6
        start_column = 1
        end_column = 34
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value:
                    cell.value = cell.value.replace('Internal','Int')
                    cell.value = cell.value.replace('External','Ext')
        workbook.save(self.output_file)
    def insert_blank_row(self):
        workbook = load_workbook(self.output_file)
        worksheet = workbook.active
        worksheet.insert_rows(7, amount=1)
        workbook.save(self.output_file)
    def shift_af5_toaf6(self):
        workbook = load_workbook(self.output_file)
        worksheet = workbook.active
        worksheet.move_range("AF5:AF6", rows=1, cols=0)
        worksheet.move_range("AG5:AG6", rows=1, cols=0)
        workbook.save(self.output_file)
    def make_column_bold(self):
        workbook = load_workbook(self.output_file)
        worksheet = workbook.active
        start_row = 6
        end_row = 118
        start_column = 4
        end_column = 4
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = Font(bold=True)
        workbook.save(self.output_file)
    def make_seventh_row_blank(self):
        workbook = load_workbook(self.output_file)
        worksheet = workbook.active
        start_row = 7
        end_row = 7
        start_column = 1
        end_column = 34
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.value = ''
        workbook.save(self.output_file)
    def change_font_to_Times_New_Roman(self):
        workbook = load_workbook(self.output_file)
        worksheet = workbook.active
        start_row = 1
        end_row = 118
        start_column = 1
        end_column = 35
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.font.bold:
                    cell.font = Font(name='Times New Roman', bold=True)
                else:
                    cell.font = Font(name='Times New Roman')
        workbook.save(self.output_file)
    def make_5th_row_wrap_text(self):
        workbook = load_workbook(self.output_file)
        worksheet = workbook.active
        start_row = 5
        end_row = 5
        start_column = 1
        end_column = 35
        for row in range(start_row, end_row + 1):
            for col in range(start_column, end_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = Font(name='Times New Roman', bold=True)
                cell.alignment = Alignment(wrap_text=True)
        workbook.save(self.output_file)
# Usage
processor = ResultProcessor('input.csv', 'output.xlsx')
processor.read_data()
processor.rename_columns()
processor.calculate_total()
processor.calculate_cgpa()
processor.process_reappear()
processor.process_absents()
processor.update_reappear_absent_columns()
processor.final_rename_columns()
processor.save_result()
processor.add_footers_to_excel()
processor.add_headers_to_excel()
processor.merge_column_names()
processor.replace()
processor.insert_blank_row()
processor.shift_af5_toaf6()
processor.add_borders_to_data()
processor.make_column_bold()
processor.make_seventh_row_blank()
processor.change_font_to_Times_New_Roman()
processor.set_height_fifth_row()  
processor.styling_sixth_row()
processor.make_5th_row_wrap_text()